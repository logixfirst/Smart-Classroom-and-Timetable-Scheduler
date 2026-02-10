"""
Attendance Management Views with RBAC
Handles Student, Faculty, and Admin operations
"""
import csv
import io

from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .attendance_models import (
    AttendanceAlert,
    AttendanceAuditLog,
    AttendanceRecord,
    AttendanceSession,
    AttendanceThreshold,
    SubjectEnrollment,
)
from .attendance_serializers import (
    AttendanceImportSerializer,
    AttendanceRecordSerializer,
    AttendanceSessionDetailSerializer,
    AttendanceSessionSerializer,
    BulkAttendanceMarkSerializer,
    StudentAttendanceSummarySerializer,
)
from .models import Faculty, Student, Subject


def get_client_ip(request):
    """Get client IP address"""
    x_forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if x_forwarded_for:
        ip = x_forwarded_for.split(",")[0]
    else:
        ip = request.META.get("REMOTE_ADDR")
    return ip


class AttendanceSessionViewSet(viewsets.ModelViewSet):
    """
    Attendance Session Management
    Faculty: Create and manage their sessions
    Admin: Full access
    Student: Read-only for their enrolled subjects
    """

    serializer_class = AttendanceSessionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user

        if user.role.lower() == "admin":
            return AttendanceSession.objects.all().select_related(
                "subject", "faculty", "batch"
            )

        elif user.role.lower() == "faculty":
            # Faculty can see their own sessions
            faculty = Faculty.objects.get(email=user.email)
            return AttendanceSession.objects.filter(faculty=faculty).select_related(
                "subject", "faculty", "batch"
            )

        elif user.role.lower() == "student":
            # Students can see sessions of their enrolled subjects
            student = Student.objects.get(email=user.email)
            enrolled_subjects = SubjectEnrollment.objects.filter(
                student=student, is_active=True
            ).values_list("subject_id", flat=True)

            return AttendanceSession.objects.filter(
                subject_id__in=enrolled_subjects
            ).select_related("subject", "faculty", "batch")

        return AttendanceSession.objects.none()

    def get_serializer_class(self):
        if self.action == "retrieve" or self.action == "mark_attendance":
            return AttendanceSessionDetailSerializer
        return AttendanceSessionSerializer

    def perform_create(self, serializer):
        # Only faculty and admin can create sessions
        if self.request.user.role not in ["faculty", "admin"]:
            raise PermissionError("Only faculty and admin can create sessions")

        serializer.save()

    @action(detail=True, methods=["post"], url_path="mark-attendance")
    def mark_attendance(self, request, pk=None):
        """
        Faculty marks attendance for a session
        POST /api/attendance/sessions/{id}/mark-attendance/
        Body: {"attendance_data": [{"student_id": "S001", "status": "present"}, ...]}
        """
        session = self.get_object()
        user = request.user

        # Permission check
        if user.role not in ["faculty", "admin"]:
            return Response(
                {"error": "Only faculty and admin can mark attendance"},
                status=status.HTTP_403_FORBIDDEN,
            )

        if user.role.lower() == "faculty":
            faculty = Faculty.objects.get(email=user.email)
            if session.faculty != faculty:
                return Response(
                    {"error": "You can only mark attendance for your own sessions"},
                    status=status.HTTP_403_FORBIDDEN,
                )

        serializer = BulkAttendanceMarkSerializer(
            data={
                "session_id": session.session_id,
                "attendance_data": request.data.get("attendance_data", []),
            }
        )

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Mark attendance
        marked_count = 0
        updated_count = 0
        errors = []

        faculty = (
            Faculty.objects.get(email=user.email) if user.role.lower() == "faculty" else None
        )

        for record_data in serializer.validated_data["attendance_data"]:
            try:
                student = Student.objects.get(student_id=record_data["student_id"])

                # Create or update attendance record
                record, created = AttendanceRecord.objects.update_or_create(
                    session=session,
                    student=student,
                    defaults={
                        "status": record_data["status"],
                        "marked_by": faculty,
                        "remarks": record_data.get("remarks", ""),
                        "verification_method": "manual",
                    },
                )

                # Create audit log
                AttendanceAuditLog.objects.create(
                    record=record,
                    action="create" if created else "update",
                    old_status=None if created else record.status,
                    new_status=record_data["status"],
                    changed_by=user,
                    reason=f"Attendance marking by {user.username}",
                    ip_address=get_client_ip(request),
                )

                if created:
                    marked_count += 1
                else:
                    updated_count += 1

            except Student.DoesNotExist:
                errors.append(f"Student {record_data['student_id']} not found")
            except Exception as e:
                errors.append(f"Error marking {record_data['student_id']}: {str(e)}")

        # Mark session as marked
        session.is_marked = True
        session.marked_at = timezone.now()
        session.save()

        # Check for low attendance and create alerts
        self._check_and_create_alerts(session)

        return Response(
            {
                "message": "Attendance marked successfully",
                "marked": marked_count,
                "updated": updated_count,
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )

    def _check_and_create_alerts(self, session):
        """Check attendance percentages and create alerts for at-risk students"""
        # Get threshold
        threshold = AttendanceThreshold.objects.filter(
            Q(department=session.subject.department) | Q(course__isnull=True),
            is_active=True,
        ).first()

        if not threshold:
            return

        # Get all students in this session
        students = Student.objects.filter(
            course=session.batch.course,
            year=session.batch.year,
            semester=session.batch.semester,
        )

        for student in students:
            # Calculate attendance percentage for this subject
            total_sessions = AttendanceRecord.objects.filter(
                student=student, session__subject=session.subject
            ).count()

            if total_sessions == 0:
                continue

            attended = AttendanceRecord.objects.filter(
                student=student,
                session__subject=session.subject,
                status__in=["present", "late"],
            ).count()

            percentage = (attended / total_sessions) * 100

            # Create alert if below threshold
            if percentage < threshold.minimum_percentage:
                AttendanceAlert.objects.create(
                    alert_type="low_attendance",
                    student=student,
                    subject=session.subject,
                    message=(
                        f"Attendance for {session.subject.subject_name} is "
                        f"{percentage:.1f}% (below {threshold.minimum_percentage}%)"
                    ),
                    severity="high" if percentage < 60 else "medium",
                )

    @action(detail=False, methods=["get"], url_path="pending")
    def pending_sessions(self, request):
        """
        Get sessions pending attendance marking
        GET /api/attendance/sessions/pending/
        """
        user = request.user

        if user.role.lower() == "faculty":
            faculty = Faculty.objects.get(email=user.email)
            sessions = AttendanceSession.objects.filter(
                faculty=faculty, is_marked=False, date__lte=timezone.now().date()
            ).select_related("subject", "batch")

        elif user.role.lower() == "admin":
            sessions = AttendanceSession.objects.filter(
                is_marked=False, date__lte=timezone.now().date()
            ).select_related("subject", "faculty", "batch")
        else:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        serializer = self.get_serializer(sessions, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="import-attendance")
    def import_attendance(self, request, pk=None):
        """
        Import attendance from CSV/Excel file
        POST /api/attendance/sessions/{id}/import-attendance/
        """
        session = self.get_object()
        user = request.user

        # Permission check
        if user.role not in ["faculty", "admin"]:
            return Response(
                {"error": "Only faculty and admin can import attendance"},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = AttendanceImportSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        file = serializer.validated_data["file"]

        try:
            if file.name.endswith(".csv"):
                return self._import_from_csv(file, session, user, request)
            else:
                return self._import_from_excel(file, session, user, request)
        except Exception as e:
            return Response(
                {"error": f"Import failed: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

    def _import_from_csv(self, file, session, user, request):
        """Import attendance from CSV"""
        decoded_file = file.read().decode("utf-8")
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)

        marked_count = 0
        errors = []
        faculty = (
            Faculty.objects.get(email=user.email) if user.role.lower() == "faculty" else None
        )

        for row in reader:
            try:
                student = Student.objects.get(student_id=row["student_id"])
                status_value = row["status"].lower()

                if status_value not in ["present", "absent", "late", "excused"]:
                    errors.append(
                        f"Invalid status for {row['student_id']}: {status_value}"
                    )
                    continue

                record, created = AttendanceRecord.objects.update_or_create(
                    session=session,
                    student=student,
                    defaults={
                        "status": status_value,
                        "marked_by": faculty,
                        "remarks": row.get("remarks", ""),
                        "verification_method": "manual",
                    },
                )

                AttendanceAuditLog.objects.create(
                    record=record,
                    action="create" if created else "update",
                    old_status=None if created else record.status,
                    new_status=status_value,
                    changed_by=user,
                    reason="Bulk import from CSV",
                    ip_address=get_client_ip(request),
                )

                marked_count += 1

            except Student.DoesNotExist:
                errors.append(f"Student not found: {row.get('student_id', 'unknown')}")
            except Exception as e:
                errors.append(f"Error: {str(e)}")

        session.is_marked = True
        session.marked_at = timezone.now()
        session.save()

        return Response(
            {"message": "Import completed", "imported": marked_count, "errors": errors}
        )

    def _import_from_excel(self, file, session, user, request):
        """Import attendance from Excel"""
        workbook = load_workbook(file)
        sheet = workbook.active

        marked_count = 0
        errors = []
        faculty = (
            Faculty.objects.get(email=user.email) if user.role.lower() == "faculty" else None
        )

        # Skip header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                student_id, status_value, remarks = (
                    row[0],
                    row[1],
                    row[2] if len(row) > 2 else "",
                )

                student = Student.objects.get(student_id=student_id)
                status_value = status_value.lower()

                if status_value not in ["present", "absent", "late", "excused"]:
                    errors.append(f"Invalid status for {student_id}: {status_value}")
                    continue

                record, created = AttendanceRecord.objects.update_or_create(
                    session=session,
                    student=student,
                    defaults={
                        "status": status_value,
                        "marked_by": faculty,
                        "remarks": remarks or "",
                        "verification_method": "manual",
                    },
                )

                AttendanceAuditLog.objects.create(
                    record=record,
                    action="create" if created else "update",
                    old_status=None if created else record.status,
                    new_status=status_value,
                    changed_by=user,
                    reason="Bulk import from Excel",
                    ip_address=get_client_ip(request),
                )

                marked_count += 1

            except Student.DoesNotExist:
                errors.append(f"Student not found: {student_id}")
            except Exception as e:
                errors.append(f"Error: {str(e)}")

        session.is_marked = True
        session.marked_at = timezone.now()
        session.save()

        return Response(
            {"message": "Import completed", "imported": marked_count, "errors": errors}
        )


class StudentAttendanceViewSet(viewsets.ViewSet):
    """
    Student Attendance Views
    Students can only view their own attendance
    """

    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=["get"], url_path="my-attendance")
    def my_attendance(self, request):
        """
        Get student's own attendance summary
        GET /api/attendance/students/my-attendance/
        """
        user = request.user

        if user.role != "student":
            return Response(
                {"error": "This endpoint is only for students"},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            student = Student.objects.get(email=user.email)
        except Student.DoesNotExist:
            return Response(
                {"error": "Student profile not found"}, status=status.HTTP_404_NOT_FOUND
            )

        # Get attendance summary
        summary = self._calculate_student_attendance(student)

        serializer = StudentAttendanceSummarySerializer(summary)
        return Response(serializer.data)

    @action(
        detail=False, methods=["get"], url_path="my-attendance/(?P<subject_id>[^/.]+)"
    )
    def my_subject_attendance(self, request, subject_id=None):
        """
        Get detailed attendance for a specific subject
        GET /api/attendance/students/my-attendance/{subject_id}/
        """
        user = request.user

        if user.role.lower() != "student":
            return Response(
                {"error": "This endpoint is only for students"},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            student = Student.objects.get(email=user.email)
            subject = Subject.objects.get(subject_id=subject_id)
        except (Student.DoesNotExist, Subject.DoesNotExist):
            return Response(
                {"error": "Student or Subject not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Get all attendance records for this subject
        records = (
            AttendanceRecord.objects.filter(student=student, session__subject=subject)
            .select_related("session", "marked_by")
            .order_by("-session__date")
        )

        serializer = AttendanceRecordSerializer(records, many=True)

        # Calculate statistics
        total = records.count()
        present = records.filter(Q(status="present") | Q(status="late")).count()
        absent = records.filter(status="absent").count()
        excused = records.filter(status="excused").count()
        percentage = (present / total * 100) if total > 0 else 0

        return Response(
            {
                "subject": {"id": subject.subject_id, "name": subject.subject_name},
                "statistics": {
                    "total_classes": total,
                    "present": present,
                    "absent": absent,
                    "excused": excused,
                    "percentage": round(percentage, 2),
                },
                "records": serializer.data,
            }
        )

    def _calculate_student_attendance(self, student):
        """Calculate comprehensive attendance summary for a student"""
        # Get all enrolled subjects
        enrollments = SubjectEnrollment.objects.filter(
            student=student, is_active=True
        ).select_related("subject")

        subject_wise = []
        total_classes = 0
        total_attended = 0
        total_missed = 0
        total_late = 0
        total_excused = 0

        # Get threshold
        threshold = AttendanceThreshold.objects.filter(
            Q(department=student.department) | Q(course__isnull=True), is_active=True
        ).first()
        min_percentage = threshold.minimum_percentage if threshold else 75.0

        for enrollment in enrollments:
            records = AttendanceRecord.objects.filter(
                student=student, session__subject=enrollment.subject
            ).select_related('session', 'marked_by').order_by('-session__date')

            total = records.count()
            present = records.filter(Q(status="present") | Q(status="late")).count()
            absent = records.filter(status="absent").count()
            late = records.filter(status="late").count()
            excused = records.filter(status="excused").count()

            percentage = (present / total * 100) if total > 0 else 0

            # Get recent records for this subject
            recent_records = []
            for record in records[:20]:  # Last 20 records
                recent_records.append({
                    "id": record.record_id,
                    "session_id": record.session.session_id,
                    "date": str(record.session.date),
                    "subject_name": enrollment.subject.subject_name,
                    "status": record.status,
                    "marked_by_name": record.marked_by.faculty_name if record.marked_by else "System",
                    "marked_at": record.marked_at.isoformat() if record.marked_at else None,
                    "remarks": record.remarks or ""
                })

            subject_wise.append(
                {
                    "subject_id": enrollment.subject.subject_id,
                    "subject_name": enrollment.subject.subject_name,
                    "total_sessions": total,
                    "present_count": present,
                    "absent_count": absent,
                    "late_count": late,
                    "excused_count": excused,
                    "attendance_percentage": round(percentage, 2),
                    "at_risk": percentage < min_percentage,
                    "recent_records": recent_records
                }
            )

            total_classes += total
            total_attended += present
            total_missed += absent
            total_late += late
            total_excused += excused

        overall_percentage = (
            (total_attended / total_classes * 100) if total_classes > 0 else 0
        )

        return {
            "student_id": student.student_id,
            "student_name": student.name,
            "subject_wise_attendance": subject_wise,
            "overall_percentage": round(overall_percentage, 2),
            "total_sessions": total_classes,
            "present_count": total_attended,
            "absent_count": total_missed,
            "late_count": total_late,
            "excused_count": total_excused,
            "at_risk": overall_percentage < min_percentage,
        }


# Continued in next part due to length...
